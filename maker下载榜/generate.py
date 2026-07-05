"""
TapTap Maker 周下载排行 HTML 生成器
用法: python generate.py
输入: 当前目录下的 20260705-maker下载榜.xlsx（可修改 INPUT_FILE）
输出: maker_download_ranking.html
"""

import re
import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
INPUT_FILE = ROOT / "20260705-maker下载榜.xlsx"
OUTPUT_FILE = ROOT.parent / "maker_download_ranking.html"


# ── 读取 Excel ────────────────────────────────────────────
def read_excel(path: Path) -> list[dict]:
    """解析 Excel，返回按周分组的数据列表，每周一个 dict。"""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=11):
        cells = [str(c.value).strip() if c.value is not None else "" for c in row]
        # 补齐到 11 列
        while len(cells) < 11:
            cells.append("")
        rows.append(cells)

    title = rows[0][0] if rows else ""

    weeks: list[dict] = []
    date_pattern = re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}\s+\d{1,2}:\d{2}(:\d{2})?$")

    i = 0
    while i < len(rows):
        row = rows[i]
        # 检测日期行：第 2 列和第 7 列都是日期（或第 2 列和第 6 列）
        start_date = ""
        end_date = ""
        for col_candidates in [(1, 5), (2, 7), (1, 6), (2, 6)]:
            ci, cj = col_candidates
            if ci < len(row) and cj < len(row):
                if date_pattern.match(row[ci]) and date_pattern.match(row[cj]):
                    start_date = row[ci]
                    end_date = row[cj]
                    break

        if start_date and end_date:
            # 下一行是表头，再往下 10 行是数据
            games: list[dict] = []
            data_start = i + 2  # 跳过表头行
            for r in range(data_start, min(data_start + 10, len(rows))):
                d = rows[r]
                if not d[0] or d[0] == "":
                    break
                try:
                    games.append({
                        "name": d[0],
                        "start_dl":    float(d[1]) if d[1] else 0,
                        "start_follow": float(d[2]) if d[2] else 0,
                        "start_rating": float(d[3]) if d[3] else 0,
                        "start_review": int(float(d[4])) if d[4] else 0,
                        "end_dl":      float(d[5]) if d[5] else 0,
                        "end_follow":   float(d[6]) if d[6] else 0,
                        "end_rating":   float(d[7]) if d[7] else 0,
                        "end_review":   int(float(d[8])) if d[8] else 0,
                        "daily_dl":     float(d[9]) if d[9] else 0,
                        "daily_review": int(float(d[10])) if d[10] else 0,
                    })
                except (ValueError, IndexError):
                    continue

            if games:
                weeks.append({
                    "start_date": start_date,
                    "end_date": end_date,
                    "games": games,
                })

            i = data_start + len(games) + 1  # 跳过数据行 + 空行
        else:
            i += 1

    wb.close()
    return title, weeks


def fmt_date(d: str) -> str:
    """2026-06-28 03:00:00 -> 2026/06/28 03:00"""
    d = d.strip()
    # 去掉秒部分
    d = re.sub(r":\d{2}$", "", d)
    # 横线转斜线
    d = d.replace("-", "/")
    return d


# ── 辅助函数 ──────────────────────────────────────────────
def fmt_num(v, decimals=2):
    """格式化数字，保留指定位小数"""
    if isinstance(v, float):
        return f"{v:.{decimals}f}"
    return str(v)


def fmt_int(v):
    return str(int(v)) if isinstance(v, float) else str(v)


def medal_html(rank: int) -> str:
    if rank == 1:
        return '<span class="medal medal-1">1</span>'
    elif rank == 2:
        return '<span class="medal medal-2">2</span>'
    elif rank == 3:
        return '<span class="medal medal-3">3</span>'
    return str(rank)


# ── HTML 模板 ──────────────────────────────────────────────
CSS = """\
:root {
    --bg: #f3f4f6;
    --surface: #ffffff;
    --surface-alt: #f8fafc;
    --text: #111827;
    --text-secondary: #6b7280;
    --text-muted: #9ca3af;
    --border: #e5e7eb;
    --border-light: #f3f4f6;
    --accent: #4f46e5;
    --accent-light: #eef2ff;
    --accent-subtle: #f5f3ff;
}
@media (prefers-color-scheme: dark) {
    :root {
        --bg: #0f1117;
        --surface: #1a1b1e;
        --surface-alt: #1f2025;
        --text: #f1f5f9;
        --text-secondary: #94a3b8;
        --text-muted: #64748b;
        --border: #2d2f34;
        --border-light: #232428;
        --accent: #818cf8;
        --accent-light: #1e1b4b;
        --accent-subtle: #1a1830;
    }
}

* { box-sizing: border-box; margin: 0; padding: 0; }

body {
    background: var(--bg);
    font-family: 'Inter', 'Noto Sans SC', system-ui, -apple-system, sans-serif;
    color: var(--text);
    padding: 48px 20px 80px;
    display: flex;
    flex-direction: column;
    align-items: center;
    min-height: 100vh;
}

.container {
    width: 100%;
    max-width: 1080px;
    display: flex;
    flex-direction: column;
    gap: 32px;
}

.page-header {
    text-align: center;
    padding-bottom: 4px;
}
.page-header h1 {
    font-size: 28px;
    font-weight: 700;
    letter-spacing: -0.3px;
    color: var(--text);
    margin-bottom: 6px;
}
.page-header .meta {
    font-size: 14px;
    color: var(--text-secondary);
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 16px;
    flex-wrap: wrap;
}
.page-header .badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 12px;
    font-weight: 600;
    background: var(--accent-light);
    color: var(--accent);
}

.week-section {
    background: var(--surface);
    border-radius: 12px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04), 0 1px 2px rgba(0,0,0,0.03);
    overflow: hidden;
    border: 1px solid var(--border);
}
.week-section.latest {
    border-color: var(--accent);
    box-shadow: 0 0 0 2px rgba(79,70,229,0.10), 0 4px 16px rgba(0,0,0,0.06);
}

.week-head {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 18px 24px;
    border: 1px solid var(--border);
    background: var(--surface-alt);
    gap: 12px;
    flex-wrap: wrap;
}
.week-section.latest .week-head {
    background: var(--accent-subtle);
    border-bottom-color: var(--accent);
}
.week-head .week-label {
    font-size: 14px;
    font-weight: 700;
    color: var(--text);
    display: flex;
    align-items: center;
    gap: 10px;
}
.week-section.latest .week-label { color: var(--accent); }
.week-head .week-date {
    font-size: 13px;
    color: var(--text-secondary);
    font-weight: 400;
}
.week-head .arrow { color: var(--text-muted); font-size: 12px; }
.week-head .new-tag {
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    background: var(--accent);
    color: #fff;
    padding: 2px 8px;
    border-radius: 4px;
}
.week-head .count {
    font-size: 12px;
    color: var(--text-muted);
    background: var(--surface);
    padding: 3px 10px;
    border-radius: 12px;
    border: 1px solid var(--border);
}

.table-wrap { overflow-x: auto; padding: 4px; }
table {
    width: 100%;
    border-collapse: collapse;
    font-size: 13.5px;
    font-variant-numeric: tabular-nums;
    white-space: nowrap;
    min-width: 920px;
    border: 2px solid var(--border);
}
thead th {
    padding: 10px 12px;
    font-weight: 600;
    font-size: 11.5px;
    text-transform: uppercase;
    letter-spacing: 0.3px;
    color: var(--text-secondary);
    border: 1px solid var(--border);
    text-align: center;
    background: var(--surface-alt);
    vertical-align: bottom;
}
thead th.group-hdr {
    font-size: 12.5px;
    font-weight: 700;
    color: var(--text);
    letter-spacing: normal;
    text-transform: none;

}
thead th.group-start { border-left: 2px solid var(--border); }
thead th.group-divider { border-left: 2px solid var(--border); }

tbody td {
    padding: 10px 12px;
    border: 1px solid var(--border);
    text-align: center;
    vertical-align: middle;
}
tbody tr:hover td { background: var(--accent-subtle); }
tbody td.name {
    text-align: left;
    font-weight: 600;
    color: var(--text);
    max-width: 200px;
    overflow: hidden;
    text-overflow: ellipsis;
}
tbody td.rank { font-weight: 700; font-size: 13px; width: 32px; text-align: center; }
.medal {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 24px;
    height: 24px;
    border-radius: 50%;
    font-size: 12px;
    font-weight: 700;
    line-height: 1;
}
.medal-1 { background: #fef3c7; color: #92400e; }
.medal-2 { background: #f1f5f9; color: #475569; }
.medal-3 { background: #fff7ed; color: #c2410c; }
@media (prefers-color-scheme: dark) {
    .medal-1 { background: #422006; color: #fbbf24; }
    .medal-2 { background: #1e293b; color: #94a3b8; }
    .medal-3 { background: #431407; color: #fdba74; }
}
td.val-primary { font-weight: 700; color: var(--accent); font-size: 14px; }

.note {
    padding: 16px 20px;
    background: var(--surface);
    border-left: 4px solid var(--accent);
    border-radius: 0 8px 8px 0;
    color: var(--text-secondary);
    font-size: 13px;
    line-height: 1.7;
    border: 1px solid var(--border);
    border-left-width: 4px;
}
.note strong { color: var(--text); }
.page-footer {
    text-align: center;
    font-size: 12px;
    color: var(--text-muted);
    padding-top: 4px;
}
"""


def build_table(week: dict) -> str:
    """为单周数据构建 <table> HTML"""
    start = fmt_date(week["start_date"])
    end = fmt_date(week["end_date"])
    games = week["games"]

    rows_html = ""
    for idx, g in enumerate(games):
        rank = idx + 1
        daily_dl = fmt_num(g["daily_dl"], 3)
        daily_dl_class = ' class="val-primary"' if rank <= 3 else ""
        rows_html += f"""\
                <tr>
                    <td class="rank">{medal_html(rank)}</td>
                    <td class="name">{g["name"]}</td>
                    <td>{fmt_num(g["start_dl"], 2)}</td>
                    <td>{fmt_num(g["start_follow"], 2)}</td>
                    <td>{fmt_num(g["start_rating"], 1)}</td>
                    <td>{fmt_int(g["start_review"])}</td>
                    <td>{fmt_num(g["end_dl"], 2)}</td>
                    <td>{fmt_num(g["end_follow"], 2)}</td>
                    <td>{fmt_num(g["end_rating"], 1)}</td>
                    <td>{fmt_int(g["end_review"])}</td>
                    <td{daily_dl_class}>{daily_dl}</td>
                    <td>{fmt_int(g["daily_review"])}</td>
                </tr>
"""

    return f"""\
        <div class="table-wrap"><table>
            <thead>
                <tr>
                    <th></th><th></th>
                    <th class="group-hdr group-start" colspan="4">{start}（期初）</th>
                    <th class="group-hdr group-divider" colspan="4">{end}（期末）</th>
                    <th class="group-hdr group-divider" colspan="2">日均新增</th>
                </tr>
                <tr>
                    <th>#</th><th>游戏名称</th>
                    <th>下载量<br>(万)</th><th>关注量<br>(万)</th><th>评分</th><th>评价<br>(个)</th>
                    <th>下载量<br>(万)</th><th>关注量<br>(万)</th><th>评分</th><th>评价<br>(个)</th>
                    <th>下载<br>(万)</th><th>评价<br>(个)</th>
                </tr>
            </thead>
            <tbody>
{rows_html}
            </tbody>
        </table></div>"""


def build_page(title: str, weeks: list[dict]) -> str:
    """构建完整 HTML 页面"""
    # 按日期排序（虽然 Excel 已按时间排列，但确保最新在前）
    # weeks 已经是从 Excel 读取的顺序（旧->新），反转使最新在前
    ordered = list(reversed(weeks))

    # 统计信息
    total_weeks = len(ordered)
    if ordered:
        first_start = fmt_date(ordered[-1]["start_date"])[:10]  # 最早一周的期初
        last_end = fmt_date(ordered[0]["end_date"])[:10]         # 最新一周的期末
        date_range = f"{first_start} – {last_end}"
    else:
        date_range = ""

    week_blocks = ""
    for i, w in enumerate(ordered):
        is_latest = (i == 0)
        week_num = total_weeks - i
        latest_class = " latest" if is_latest else ""
        new_tag = '<span class="new-tag">NEW</span> ' if is_latest else ""
        week_blocks += f"""\
    <!-- ── Week {week_num} ── -->
    <div class="week-section{latest_class}">
        <div class="week-head">
            <div class="week-label">{new_tag}第 {week_num} 周</div>
            <span class="week-date">{fmt_date(w["start_date"])} <span class="arrow">→</span> {fmt_date(w["end_date"])}</span>
            <span class="count">Top 10</span>
        </div>
{build_table(w)}
    </div>

"""

    today_str = datetime.date.today().isoformat()

    return f"""\
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title} — Maker 下载榜</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Noto+Sans+SC:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
{CSS}
    </style>
</head>
<body>

<div class="container">

    <!-- Header -->
    <div class="page-header">
        <h1>{title}</h1>
        <div class="meta">
            <span>📊 Maker 游戏周度追踪</span>
            <span>📅 {date_range}</span>
            <span class="badge">共 {total_weeks} 周</span>
        </div>
    </div>

{week_blocks}
    <div class="note">
        <strong>💡 数据说明：</strong>每周选取 PC 端新增下载量前 10 的 Maker 游戏，按日均新增下载（万）降序排列。日均新增 =（期末数据 − 期初数据）÷ 间隔天数。数据来源：TapTap 爬虫，每周末定时抓取。
    </div>

    <p class="page-footer">Generated on {today_str} · TapTap 数据分析大盘</p>

</div>

</body>
</html>
"""


# ── Main ──────────────────────────────────────────────────
def main():
    if not INPUT_FILE.exists():
        print(f"[ERROR] 找不到输入文件: {INPUT_FILE}")
        print("   请将 Excel 文件放到脚本同目录下，或修改 INPUT_FILE 变量。")
        return

    print(f"[READ] 读取: {INPUT_FILE.name}")
    title, weeks = read_excel(INPUT_FILE)

    if not weeks:
        print("[ERROR] 未能解析出任何周数据，请检查 Excel 格式。")
        return

    print(f"   -> 解析到 {len(weeks)} 周数据")

    html = build_page(title, weeks)

    OUTPUT_FILE.write_text(html, encoding="utf-8")
    print(f"[DONE] 已生成: {OUTPUT_FILE.name}  ({len(html):,} bytes)")


if __name__ == "__main__":
    main()
